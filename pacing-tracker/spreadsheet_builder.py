"""Build a WTD performance-to-goal Excel workbook from campaign CSV data.

Scans `pacing-tracker/reports/` for CSVs, aggregates spend, budget goal,
impressions, and clicks per platform bucket + campaign + publisher + tactic
(with `dv360 + youtube` split into its own bucket), and writes a formatted
`wtd_performance_to_goal.xlsx` workbook with one tab per CSV.
"""

import argparse
import csv
from collections import Counter, defaultdict
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DEFAULT_REPORTS_DIR = "reports"
DEFAULT_OUTPUT_FILE = "wtd_performance_to_goal.xlsx"

SLACK_UNASSOCIATED_FILE = "slack_unassociated.txt"
SLACK_PREVIEW_FILE = "slack_preview.txt"
LEGACY_MISSING_FILES = ("performance_missing_goal.csv", "goals_missing_performance.csv")
RESERVED_OUTPUT_NAMES = {SLACK_UNASSOCIATED_FILE, SLACK_PREVIEW_FILE, *LEGACY_MISSING_FILES}

MISSING_GOAL_HEADERS = ["CAMPAIGN", "SPEND", "IMPRESSIONS", "CLICKS", "PUBLISHER", "PLATFORM"]
MISSING_PERF_HEADERS = ["CAMPAIGN", "PLATFORM", "PUBLISHER", "SPEND GOAL"]

SLACK_UNASSOCIATED_HEADER = (
    "The pacing tracker was unable to associate campaign metrics and budgets "
    "for the following campaigns:"
)

NULL_BUCKET = "null"

INVALID_SHEET_CHARS = set(r"\/?*[]:")
SHEET_NAME_LIMIT = 31

# Visual styling -------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

TITLE_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FONT = Font(bold=True, size=11, color="1F3864")

LABEL_FILL = PatternFill("solid", fgColor="F2F2F2")
LABEL_FONT = Font(bold=True, color="333333")

SECTION_FILL = PatternFill("solid", fgColor="BDD7EE")
SECTION_FONT = Font(bold=True, size=11, color="1F3864")

SECTION_TOTAL_FILL = PatternFill("solid", fgColor="8EA9DB")
SECTION_TOTAL_FONT = Font(bold=True, color="FFFFFF")

GRAND_TOTAL_FILL = PatternFill("solid", fgColor="1F3864")
GRAND_TOTAL_FONT = Font(bold=True, color="FFFFFF", size=12)

NULL_SECTION_FILL = PatternFill("solid", fgColor="F8CBAD")
NULL_SECTION_FONT = Font(bold=True, size=11, color="833C0C")

THIN = Side(style="thin", color="9BA4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", indent=1)
RIGHT = Alignment(horizontal="right", vertical="center")


# Parsing helpers ------------------------------------------------------------


def parse_decimal(value: str | None) -> Decimal:
    if value is None:
        return Decimal("0")
    cleaned = value.strip()
    if not cleaned or cleaned.lower() == "null":
        return Decimal("0")
    cleaned = cleaned.replace("$", "").replace(",", "")
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return Decimal("0")


def parse_int(value: str | None) -> int:
    if value is None:
        return 0
    cleaned = value.strip()
    if not cleaned or cleaned.lower() == "null":
        return 0
    try:
        return int(Decimal(cleaned))
    except InvalidOperation:
        return 0


def normalize(value: str | None) -> str:
    if value is None:
        return ""
    cleaned = value.strip()
    if cleaned.lower() == "null":
        return ""
    return cleaned


def is_null_value(value: str | None) -> bool:
    return normalize(value) == ""


def get_bucket(platform: str, publisher: str) -> str:
    if not platform:
        return NULL_BUCKET
    if platform.lower() == "dv360" and publisher.lower() == "youtube":
        return "dv360 youtube"
    return platform.lower()


# Aggregation ----------------------------------------------------------------


def aggregate_rows(csv_path: Path):
    """Return (buckets, retail_week, missing_goal_rows, missing_perf_rows).

    buckets: dict[bucket_name, dict] with keys:
        - details: list[dict] of aggregated detail rows
        - totals: dict of section totals
    retail_week: most common retail_week value in the CSV (string) or "".
    missing_goal_rows: list[dict] for performance_missing_goal.csv (rows with
        null update_total_budget).
    missing_perf_rows: list[dict] for goals_missing_performance.csv (rows with
        null spend).
    """

    # detail_key -> bucket name -> aggregated metrics
    bucket_details: dict[str, dict[tuple[str, str, str], dict]] = defaultdict(dict)
    bucket_totals: dict[str, dict] = defaultdict(
        lambda: {
            "spend": Decimal("0"),
            "goal": Decimal("0"),
            "impressions": 0,
            "clicks": 0,
            "rows": 0,
        }
    )
    retail_week_counter: Counter[str] = Counter()
    missing_goal_rows: list[dict] = []
    missing_perf_rows: list[dict] = []

    with csv_path.open(mode="r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.DictReader(csv_file)
        for row in reader:
            platform = normalize(row.get("platform"))
            publisher = normalize(row.get("publisher"))
            bucket = get_bucket(platform, publisher)

            campaign = normalize(row.get("initiativecampaign")) or "(unspecified)"
            tactic = normalize(row.get("tactic")) or "(unspecified)"
            publisher_display = publisher or "(unspecified)"

            key = (campaign, publisher_display, tactic)
            detail = bucket_details[bucket].setdefault(
                key,
                {
                    "campaign": campaign,
                    "publisher": publisher_display,
                    "tactic": tactic,
                    "spend": Decimal("0"),
                    "goal": Decimal("0"),
                    "impressions": 0,
                    "clicks": 0,
                    "rows": 0,
                },
            )

            spend = parse_decimal(row.get("spend"))
            goal = parse_decimal(row.get("update_total_budget"))
            impressions = parse_int(row.get("impressions"))
            clicks = parse_int(row.get("clicks"))

            detail["spend"] += spend
            detail["goal"] += goal
            detail["impressions"] += impressions
            detail["clicks"] += clicks
            detail["rows"] += 1

            totals = bucket_totals[bucket]
            totals["spend"] += spend
            totals["goal"] += goal
            totals["impressions"] += impressions
            totals["clicks"] += clicks
            totals["rows"] += 1

            retail_week = normalize(row.get("retail_week"))
            if retail_week:
                retail_week_counter[retail_week] += 1

            if is_null_value(row.get("update_total_budget")):
                missing_goal_rows.append(
                    {
                        "campaign_name": normalize(row.get("big_boi_campaign")),
                        "spend": float(spend),
                        "impressions": impressions,
                        "clicks": clicks,
                        "publisher": normalize(row.get("harmonized_publisher_platform")),
                        "platform": normalize(row.get("harmonized_partner")),
                    }
                )

            if is_null_value(row.get("spend")):
                missing_perf_rows.append(
                    {
                        "campaign_name": normalize(row.get("initiativecampaign")),
                        "platform": platform,
                        "publisher": publisher,
                        "spend_goal": float(goal),
                    }
                )

    buckets: dict[str, dict] = {}
    for bucket, details_map in bucket_details.items():
        details = sorted(
            details_map.values(),
            key=lambda d: (d["publisher"].lower(), d["tactic"].lower(), d["campaign"].lower()),
        )
        buckets[bucket] = {"details": details, "totals": bucket_totals[bucket]}

    retail_week = (
        retail_week_counter.most_common(1)[0][0] if retail_week_counter else ""
    )
    return buckets, retail_week, missing_goal_rows, missing_perf_rows


# Week / pacing helpers ------------------------------------------------------


def percent_through_current_week() -> tuple[date, date, float]:
    """Return (start_of_week_sunday, end_of_week_saturday, pct_through_week).

    Reference date is yesterday — pacing reports run the morning after, so
    "yesterday" is the latest day with complete data. Percent is the inclusive
    day count through yesterday divided by 7.
    """

    reference = date.today() - timedelta(days=1)
    # Python weekday: Monday=0 ... Sunday=6. We want Sunday=0 ... Saturday=6.
    sunday_offset = (reference.weekday() + 1) % 7
    start = reference - timedelta(days=sunday_offset)
    end = start + timedelta(days=6)
    elapsed_days = (reference - start).days + 1  # inclusive day count
    pct = elapsed_days / 7
    return start, end, pct


# Workbook building ----------------------------------------------------------


COLUMNS: list[tuple[str, str, str]] = [
    # (header, kind, width)
    ("CAMPAIGN", "text", 28),
    ("PUBLISHER", "text", 18),
    ("TACTIC", "text", 20),
    ("TOTAL SPEND", "currency", 16),
    ("SPEND GOAL", "currency", 16),
    ("% TO GOAL", "percent", 12),
    ("IMPRESSIONS", "int", 14),
    ("CLICKS", "int", 12),
    ("CTR", "percent", 10),
]


def _fmt(kind: str) -> str:
    return {
        "currency": '"$"#,##0',
        "percent": "0.00%",
        "int": "#,##0",
        "text": "@",
    }[kind]


def _apply_row_format(
    ws, row_idx: int, fill: PatternFill | None, font: Font | None
) -> None:
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = BORDER
        if fill is not None:
            cell.fill = fill
        if font is not None:
            cell.font = font


def _write_detail_row(ws, row_idx: int, detail: dict) -> None:
    values = [
        detail["campaign"].upper(),
        detail["publisher"].upper(),
        detail["tactic"].upper(),
        float(detail["spend"]),
        float(detail["goal"]),
        None,  # % to goal formula
        detail["impressions"],
        detail["clicks"],
        None,  # CTR formula
    ]
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        kind = COLUMNS[col_idx - 1][1]
        cell.number_format = _fmt(kind)
        if kind in ("currency", "percent", "int"):
            cell.alignment = RIGHT
        elif col_idx == 1:
            cell.alignment = LEFT
        else:
            cell.alignment = CENTER

    spend_cell = f"D{row_idx}"
    goal_cell = f"E{row_idx}"
    impressions_cell = f"G{row_idx}"
    clicks_cell = f"H{row_idx}"

    ws.cell(row=row_idx, column=6).value = (
        f'=IFERROR({spend_cell}/{goal_cell},0)'
    )
    ws.cell(row=row_idx, column=9).value = (
        f'=IFERROR({clicks_cell}/{impressions_cell},0)'
    )

    _apply_row_format(ws, row_idx, fill=None, font=None)


def _write_total_row(
    ws,
    row_idx: int,
    label: str,
    start_row: int,
    end_row: int,
    *,
    fill: PatternFill,
    font: Font,
) -> None:
    ws.cell(row=row_idx, column=1, value=label)
    ws.merge_cells(start_row=row_idx, end_row=row_idx, start_column=1, end_column=3)

    ws.cell(row=row_idx, column=4, value=f"=SUM(D{start_row}:D{end_row})")
    ws.cell(row=row_idx, column=5, value=f"=SUM(E{start_row}:E{end_row})")
    ws.cell(
        row=row_idx,
        column=6,
        value=f'=IFERROR(D{row_idx}/E{row_idx},0)',
    )
    ws.cell(row=row_idx, column=7, value=f"=SUM(G{start_row}:G{end_row})")
    ws.cell(row=row_idx, column=8, value=f"=SUM(H{start_row}:H{end_row})")
    ws.cell(
        row=row_idx,
        column=9,
        value=f'=IFERROR(H{row_idx}/G{row_idx},0)',
    )

    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        kind = COLUMNS[col_idx - 1][1]
        cell.number_format = _fmt(kind)
        cell.fill = fill
        cell.font = font
        cell.border = BORDER
        if col_idx == 1:
            cell.alignment = LEFT
        elif kind in ("currency", "percent", "int"):
            cell.alignment = RIGHT
        else:
            cell.alignment = CENTER


def _write_section_header(ws, row_idx: int, label: str, is_null: bool) -> None:
    ws.cell(row=row_idx, column=1, value=label.upper())
    ws.merge_cells(
        start_row=row_idx,
        end_row=row_idx,
        start_column=1,
        end_column=len(COLUMNS),
    )
    fill = NULL_SECTION_FILL if is_null else SECTION_FILL
    font = NULL_SECTION_FONT if is_null else SECTION_FONT
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.border = BORDER
        cell.alignment = LEFT


def _write_top_section(ws, retail_week: str) -> int:
    start, end, pct = percent_through_current_week()

    ws.cell(row=2, column=1, value="RETAIL WEEK")
    ws.cell(row=2, column=2, value=retail_week or "N/A")

    ws.cell(row=3, column=1, value="WEEK START (SUN)")
    ws.cell(row=3, column=2, value=start.isoformat())

    ws.cell(row=4, column=1, value="WEEK END (SAT)")
    ws.cell(row=4, column=2, value=end.isoformat())

    ws.cell(row=5, column=1, value="% THROUGH WEEK")
    pct_cell = ws.cell(row=5, column=2, value=pct)
    pct_cell.number_format = "0.00%"

    for row in range(2, 6):
        label_cell = ws.cell(row=row, column=1)
        value_cell = ws.cell(row=row, column=2)
        label_cell.fill = LABEL_FILL
        label_cell.font = LABEL_FONT
        label_cell.alignment = LEFT
        label_cell.border = BORDER
        value_cell.fill = TITLE_FILL
        value_cell.font = TITLE_FONT
        value_cell.alignment = CENTER
        value_cell.border = BORDER

    # Leave a blank padding row before the table header.
    return 7


def _write_table_header(ws, row_idx: int) -> None:
    for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _bucket_order(buckets: dict[str, dict]) -> list[str]:
    # Predictable, friendly ordering: known platforms first, then alpha, null last.
    preferred = ["google ads", "dv360", "dv360 youtube", "remerge"]
    remaining = sorted(b for b in buckets if b not in preferred and b != NULL_BUCKET)
    ordered = [b for b in preferred if b in buckets] + remaining
    if NULL_BUCKET in buckets:
        ordered.append(NULL_BUCKET)
    return ordered


def _populate_sheet(ws, buckets: dict[str, dict], retail_week: str) -> None:
    # Column A label column width for the top section.
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    table_start = _write_top_section(ws, retail_week)
    _write_table_header(ws, table_start)

    current_row = table_start + 1
    detail_ranges: list[tuple[int, int]] = []

    for bucket in _bucket_order(buckets):
        section = buckets[bucket]
        details = section["details"]

        section_label = "UNATTRIBUTED / NULL" if bucket == NULL_BUCKET else bucket
        _write_section_header(
            ws, current_row, section_label, is_null=(bucket == NULL_BUCKET)
        )
        current_row += 1

        section_start = current_row
        for detail in details:
            _write_detail_row(ws, current_row, detail)
            current_row += 1
        section_end = current_row - 1

        if section_end >= section_start:
            detail_ranges.append((section_start, section_end))
            total_label = (
                "UNATTRIBUTED / NULL TOTAL"
                if bucket == NULL_BUCKET
                else f"{bucket.upper()} TOTAL"
            )
            _write_total_row(
                ws,
                current_row,
                total_label,
                section_start,
                section_end,
                fill=SECTION_TOTAL_FILL,
                font=SECTION_TOTAL_FONT,
            )
            current_row += 1

        # Blank spacer row between sections.
        current_row += 1

    # Grand total (sum of all detail rows across all sections).
    if detail_ranges:
        grand_row = current_row
        parts_spend = ",".join(f"D{s}:D{e}" for s, e in detail_ranges)
        parts_goal = ",".join(f"E{s}:E{e}" for s, e in detail_ranges)
        parts_impr = ",".join(f"G{s}:G{e}" for s, e in detail_ranges)
        parts_clicks = ",".join(f"H{s}:H{e}" for s, e in detail_ranges)

        ws.cell(row=grand_row, column=1, value="GRAND TOTAL")
        ws.merge_cells(
            start_row=grand_row, end_row=grand_row, start_column=1, end_column=3
        )
        ws.cell(row=grand_row, column=4, value=f"=SUM({parts_spend})")
        ws.cell(row=grand_row, column=5, value=f"=SUM({parts_goal})")
        ws.cell(
            row=grand_row,
            column=6,
            value=f"=IFERROR(D{grand_row}/E{grand_row},0)",
        )
        ws.cell(row=grand_row, column=7, value=f"=SUM({parts_impr})")
        ws.cell(row=grand_row, column=8, value=f"=SUM({parts_clicks})")
        ws.cell(
            row=grand_row,
            column=9,
            value=f"=IFERROR(H{grand_row}/G{grand_row},0)",
        )

        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=grand_row, column=col_idx)
            kind = COLUMNS[col_idx - 1][1]
            cell.number_format = _fmt(kind)
            cell.fill = GRAND_TOTAL_FILL
            cell.font = GRAND_TOTAL_FONT
            cell.border = BORDER
            if col_idx == 1:
                cell.alignment = LEFT
            elif kind in ("currency", "percent", "int"):
                cell.alignment = RIGHT
            else:
                cell.alignment = CENTER

    # Freeze panes below the table header.
    ws.freeze_panes = f"A{table_start + 1}"


def discover_reports(reports_dir: Path) -> list[Path]:
    if not reports_dir.is_dir():
        raise SystemExit(f"Reports directory not found: {reports_dir}")
    csvs = sorted(
        p
        for p in reports_dir.iterdir()
        if p.suffix.lower() == ".csv" and p.name not in RESERVED_OUTPUT_NAMES
    )
    if not csvs:
        raise SystemExit(f"No CSV files in {reports_dir}")
    return csvs


def _escape_cell(value) -> str:
    return str(value).replace("|", r"\|")


def _format_markdown_table(headers: list[str], rows: list[list]) -> str:
    cells = [[_escape_cell(h) for h in headers]]
    cells.extend([_escape_cell(v) for v in row] for row in rows)
    widths = [max(len(row[i]) for row in cells) for i in range(len(headers))]

    def render(row: list[str]) -> str:
        return "| " + " | ".join(cell.ljust(widths[i]) for i, cell in enumerate(row)) + " |"

    separator = "| " + " | ".join("-" * widths[i] for i in range(len(headers))) + " |"
    return "\n".join([render(cells[0]), separator, *(render(row) for row in cells[1:])])


def _missing_goal_table(rows: list[dict]) -> str:
    table_rows = [
        [
            r["campaign_name"],
            f"${r['spend']:,.0f}",
            f"{r['impressions']:,}",
            f"{r['clicks']:,}",
            r["publisher"],
            r["platform"],
        ]
        for r in rows
    ]
    return _format_markdown_table(MISSING_GOAL_HEADERS, table_rows)


def _missing_perf_table(rows: list[dict]) -> str:
    table_rows = [
        [
            r["campaign_name"],
            r["platform"],
            r["publisher"],
            f"${r['spend_goal']:,.0f}",
        ]
        for r in rows
    ]
    return _format_markdown_table(MISSING_PERF_HEADERS, table_rows)


def build_slack_unassociated(
    sections: list[tuple[str, list[dict], list[dict]]],
) -> str | None:
    qualifying = [
        (name, mg, mp) for name, mg, mp in sections if mg and mp
    ]
    if not qualifying:
        return None

    blocks = []
    for csv_name, missing_goal, missing_perf in qualifying:
        block = "\n".join(
            [
                csv_name,
                "Campaign performance without goals:",
                "```",
                _missing_goal_table(missing_goal),
                "```",
                "",
                "Campaign spend goals without performance:",
                "```",
                _missing_perf_table(missing_perf),
                "```",
            ]
        )
        blocks.append(block)

    return SLACK_UNASSOCIATED_HEADER + "\n\n" + "\n\n".join(blocks) + "\n"


def build_slack_preview(
    sections: list[tuple[str, dict[str, dict]]],
) -> str | None:
    blocks = []
    for csv_name, buckets in sections:
        platforms = []
        total_spend = Decimal("0")
        total_goal = Decimal("0")
        for bucket in _bucket_order(buckets):
            if bucket == NULL_BUCKET:
                continue
            totals = buckets[bucket]["totals"]
            if totals["goal"] <= 0:
                continue
            platforms.append((bucket, totals["spend"], totals["goal"]))
            total_spend += totals["spend"]
            total_goal += totals["goal"]

        if not platforms:
            continue

        total_pct = float(total_spend / total_goal * 100) if total_goal else 0.0
        lines = [
            csv_name,
            (
                f"WTD spend was ${total_spend:,.0f} which was {total_pct:.2f}% "
                f"to our weekly budget goal of ${total_goal:,.0f}"
            ),
            "",
            "PLATFORM SPEND (% TO GOAL)",
        ]
        name_width = max(len(b.upper()) for b, _, _ in platforms)
        for bucket, spend, goal in platforms:
            pct = float(spend / goal * 100)
            lines.append(f"{bucket.upper().ljust(name_width)} = {pct:>6.0f}%")
        blocks.append("\n".join(lines))

    if not blocks:
        return None
    return "\n\n".join(blocks) + "\n"


def sheet_name_from_path(path: Path, taken: set[str]) -> str:
    cleaned = "".join("_" if c in INVALID_SHEET_CHARS else c for c in path.stem)
    name = cleaned[:SHEET_NAME_LIMIT] or "sheet"
    if name in taken:
        i = 2
        while True:
            suffix = f"_{i}"
            candidate = name[: SHEET_NAME_LIMIT - len(suffix)] + suffix
            if candidate not in taken:
                name = candidate
                break
            i += 1
    taken.add(name)
    return name


def build_workbook(
    reports: list[tuple[str, dict[str, dict], str]], output_path: Path
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, buckets, retail_week in reports:
        ws = wb.create_sheet(title=sheet_name)
        _populate_sheet(ws, buckets, retail_week)
    wb.save(output_path)


def main() -> None:
    now = datetime.now(ZoneInfo("America/Chicago"))
    if time(10, 0) <= now.time() <= time(10, 50):
        Path("time.txt").write_text(now.isoformat())

    parser = argparse.ArgumentParser(
        description=(
            "Aggregate weekly campaign performance data into a formatted "
            "Excel workbook that shows spend vs budget goal."
        )
    )
    parser.add_argument(
        "--reports-dir",
        default=DEFAULT_REPORTS_DIR,
        help="Directory containing campaign CSV files (default: reports/).",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=DEFAULT_OUTPUT_FILE,
        help="Output .xlsx path (default: wtd_performance_to_goal.xlsx).",
    )
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent

    reports_dir = Path(args.reports_dir)
    if not reports_dir.is_absolute():
        reports_dir = script_dir / reports_dir

    output_path = Path(args.output)
    if not output_path.is_absolute():
        output_path = script_dir / output_path

    csv_paths = discover_reports(reports_dir)
    taken: set[str] = set()
    reports: list[tuple[str, dict[str, dict], str]] = []
    unassociated_sections: list[tuple[str, list[dict], list[dict]]] = []
    preview_sections: list[tuple[str, dict[str, dict]]] = []
    for csv_path in csv_paths:
        buckets, retail_week, missing_goal, missing_perf = aggregate_rows(csv_path)
        reports.append((sheet_name_from_path(csv_path, taken), buckets, retail_week))
        unassociated_sections.append((csv_path.name, missing_goal, missing_perf))
        preview_sections.append((csv_path.name, buckets))

    build_workbook(reports, output_path)

    slack_path = reports_dir / SLACK_UNASSOCIATED_FILE
    preview_path = reports_dir / SLACK_PREVIEW_FILE
    for legacy_name in (
        SLACK_UNASSOCIATED_FILE,
        SLACK_PREVIEW_FILE,
        *LEGACY_MISSING_FILES,
    ):
        legacy_path = reports_dir / legacy_name
        if legacy_path.exists():
            legacy_path.unlink()

    body = build_slack_unassociated(unassociated_sections)
    if body is not None:
        slack_path.write_text(body, encoding="utf-8")
        print(f"Wrote {slack_path}")

    preview_body = build_slack_preview(preview_sections)
    if preview_body is not None:
        preview_path.write_text(preview_body, encoding="utf-8")
        print(f"Wrote {preview_path}")

    print(f"Wrote {output_path} ({len(reports)} sheets)")
    for sheet_name, buckets, retail_week in reports:
        print(f"  [{sheet_name}] retail_week={retail_week or 'N/A'}")
        for bucket, section in buckets.items():
            totals = section["totals"]
            spend = float(totals["spend"])
            goal = float(totals["goal"])
            pct = (spend / goal * 100) if goal else 0.0
            print(
                f"    {bucket}: rows={totals['rows']}, "
                f"spend=${spend:,.2f}, goal=${goal:,.2f}, %goal={pct:,.2f}%"
            )


if __name__ == "__main__":
    main()
